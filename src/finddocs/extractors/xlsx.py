"""Parser arkuszy kalkulacyjnych Office Open XML (xlsx, xlsm, xltx).

Kazdy wiersz arkusza trafia do osobnej sekcji, dzieki czemu wyszukiwarka moze
wskazac pojedynczy rekord tabeli zamiast calego pliku. Gdy pierwszy niepusty
wiersz wyglada na naglowek kolumn, wartosci komorek sa opisywane w postaci
"Kolumna: wartosc", co poprawia czytelnosc wynikow i trafnosc dopasowania.

Skoroszyt jest otwierany w trybie tylko do odczytu (openpyxl, read_only), zeby
ograniczyc zuzycie pamieci przy duzych plikach. Parser nie renderuje arkuszy
i nie wykonuje OCR.
"""

from __future__ import annotations

import datetime as dt
import math
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook

from finddocs.errors import (
    CorruptedFileError,
    EmptyDocumentError,
    ExtractionError,
    FindDocsError,
    PasswordProtectedError,
    UnsupportedFormatError,
)
from finddocs.extractors.base import ExtractionContext, Extractor
from finddocs.normalization.text import clean_text
from finddocs.types import ExtractedSection, ExtractionResult, SupportLevel, TextOrigin

if TYPE_CHECKING:
    from collections.abc import Iterator
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

#: Co ile wierszy sprawdzac anulowanie i limit czasu.
_CHECKPOINT_EVERY = 200

#: Separator laczacy komorki jednego wiersza.
_CELL_SEPARATOR = " | "

#: Sygnatura kontenera OLE. Skoroszyt zaszyfrowany haslem nie jest archiwum ZIP.
_OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

#: Fragmenty komunikatow openpyxl swiadczace o zabezpieczeniu haslem.
_ENCRYPTED_MARKERS = ("encrypt", "password", "protected", "hasl", "zaszyfrow")

#: Fragmenty komunikatow openpyxl swiadczace o nieobslugiwanym wariancie pliku.
_UNSUPPORTED_MARKERS = ("does not support", "not supported", "unsupported", "file format")


@dataclass(slots=True)
class _SheetOutcome:
    """Podsumowanie odczytu jednego arkusza."""

    next_order: int
    """Kolejny wolny numer porządkowy sekcji."""

    rows: int
    """Liczba sekcji z treścią (nagłówek oraz wiersze) dodanych dla arkusza."""

    failed: bool = False
    """True, gdy odczyt arkusza przerwał błąd biblioteki."""


def _has_ole_signature(path: Path) -> bool:
    """Czy plik jest kontenerem OLE, typowym dla skoroszytow zaszyfrowanych."""
    try:
        with path.open("rb") as handle:
            return handle.read(len(_OLE_SIGNATURE)) == _OLE_SIGNATURE
    except OSError:
        return False


def _format_datetime(value: dt.datetime) -> str:
    """Data w formacie ISO, z godzina tylko wtedy, gdy jest niezerowa."""
    if value.second:
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if value.hour or value.minute:
        return value.strftime("%Y-%m-%d %H:%M")
    return value.strftime("%Y-%m-%d")


def _format_time(value: dt.time) -> str:
    """Godzina w formacie HH:MM, z sekundami tylko gdy sa niezerowe."""
    if value.second:
        return value.strftime("%H:%M:%S")
    return value.strftime("%H:%M")


def _format_float(value: float) -> str:
    """Liczba zmiennoprzecinkowa bez zbednej koncowki '.0'."""
    if math.isnan(value) or math.isinf(value):
        return ""
    if value.is_integer():
        return str(int(value))
    return str(value)


def _format_decimal(value: Decimal) -> str:
    """Liczba dziesietna bez zbednych zer na koncu."""
    if not value.is_finite():
        return ""
    try:
        if value == value.to_integral_value():
            return str(value.to_integral_value())
    except (InvalidOperation, ValueError, OverflowError):
        return str(value)
    return format(value.normalize(), "f")


def format_cell(value: object) -> str:
    """Zamienia wartosc komorki na tekst. Pusta wartosc daje pusty string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "PRAWDA" if value else "FALSZ"
    if isinstance(value, dt.datetime):
        return _format_datetime(value)
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return _format_time(value)
    if isinstance(value, dt.timedelta):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _format_float(value)
    if isinstance(value, Decimal):
        return _format_decimal(value)
    return str(value).strip()


def _clean_meta(value: object) -> str | None:
    """Oczyszczony tekst metadanej albo None, gdy pole jest puste."""
    if value is None:
        return None
    cleaned = clean_text(str(value))
    return cleaned or None


def _as_datetime(value: object) -> dt.datetime | None:
    """Zwraca date modyfikacji lub utworzenia, gdy openpyxl podal poprawny typ."""
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    return None


class XlsxExtractor(Extractor):
    """Adapter skoroszytow Excel zapisanych w formacie Office Open XML."""

    name = "xlsx"
    extensions = (".xlsx", ".xlsm", ".xltx")
    mime_types = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
    )
    support_level = SupportLevel.FULL
    priority = 120

    def extract(self, path: Path, context: ExtractionContext) -> ExtractionResult:
        """Odczytuje wszystkie arkusze skoroszytu i zwraca sekcje wierszy."""
        result = ExtractionResult(parser_name=self.name, support_level=self.support_level)
        result.origin = TextOrigin.NATIVE
        result.needs_ocr = False

        workbook = self._open_workbook(path)
        order = 0
        content_rows = 0
        failed_sheets = 0
        sheet_names: list[str] = []
        try:
            self._read_properties(workbook, result)
            for worksheet in workbook.worksheets:
                context.checkpoint()
                sheet_names.append(str(worksheet.title))
                outcome = self._read_sheet(worksheet, context, result, order)
                order = outcome.next_order
                content_rows += outcome.rows
                failed_sheets += int(outcome.failed)
        finally:
            self._close(workbook)

        result.metadata.extra["arkusze"] = sheet_names
        result.metadata.extra["liczba_arkuszy"] = len(sheet_names)

        if content_rows == 0:
            if failed_sheets:
                raise CorruptedFileError(
                    "Nie udało się odczytać żadnego arkusza skoroszytu, plik może być uszkodzony.",
                    details={"plik": path.name, "arkusze": sheet_names},
                )
            raise EmptyDocumentError(
                "Skoroszyt Excel nie zawiera danych możliwych do zaindeksowania.",
                details={"plik": path.name, "arkusze": sheet_names},
            )
        return result

    # --- otwieranie pliku ------------------------------------------------------

    def _open_workbook(self, path: Path) -> Workbook:
        """Otwiera skoroszyt, tlumaczac bledy biblioteki na wyjatki FindDocs."""
        try:
            return load_workbook(path, read_only=True, data_only=True)
        except zipfile.BadZipFile as exc:
            if _has_ole_signature(path):
                raise PasswordProtectedError(
                    "Skoroszyt jest zaszyfrowany lub zabezpieczony hasłem.",
                    details={"plik": path.name},
                    cause=exc,
                ) from exc
            raise CorruptedFileError(
                "Plik xlsx jest uszkodzony: nie udało się odczytać archiwum ZIP.",
                details={"plik": path.name},
                cause=exc,
            ) from exc
        except FindDocsError:
            raise
        except Exception as exc:
            raise self._translate_open_error(exc, path) from exc

    def _translate_open_error(self, exc: Exception, path: Path) -> ExtractionError:
        """Dobiera wyjatek FindDocs na podstawie komunikatu biblioteki."""
        message = str(exc).casefold()
        details: dict[str, Any] = {"plik": path.name, "blad": type(exc).__name__}
        if any(marker in message for marker in _ENCRYPTED_MARKERS):
            return PasswordProtectedError(
                "Skoroszyt jest zabezpieczony hasłem, odczyt treści nie jest możliwy.",
                details=details,
            )
        if any(marker in message for marker in _UNSUPPORTED_MARKERS):
            return UnsupportedFormatError(
                "Ten wariant pliku Excel nie jest obsługiwany przez parser xlsx.",
                details=details,
            )
        if isinstance(exc, OSError):
            return ExtractionError(
                "Nie udało się odczytać pliku skoroszytu z dysku.",
                details=details,
            )
        return CorruptedFileError(
            "Nie udało się otworzyć skoroszytu Excel, plik może być uszkodzony.",
            details=details,
        )

    def _close(self, workbook: Workbook) -> None:
        """Zamyka skoroszyt. Blad zamkniecia nie moze przeslonic wyniku."""
        try:
            workbook.close()
        except Exception:
            return

    # --- metadane --------------------------------------------------------------

    def _read_properties(self, workbook: Workbook, result: ExtractionResult) -> None:
        """Przepisuje wlasciwosci dokumentu do metadanych wyniku."""
        try:
            props = workbook.properties
        except Exception:
            result.warnings.append("Nie udało się odczytać właściwości skoroszytu.")
            return
        metadata = result.metadata
        metadata.title = _clean_meta(getattr(props, "title", None))
        metadata.author = _clean_meta(getattr(props, "creator", None))
        metadata.subject = _clean_meta(getattr(props, "subject", None))
        metadata.keywords = _clean_meta(getattr(props, "keywords", None))
        metadata.created_at = _as_datetime(getattr(props, "created", None))
        metadata.modified_at = _as_datetime(getattr(props, "modified", None))

    # --- arkusze ---------------------------------------------------------------

    def _iter_rows(self, worksheet: Worksheet) -> Iterator[tuple[list[Any], int]]:
        """Zwraca wartosci komorek wraz z numerem wiersza zgodnym z Excelem."""
        number = 0
        for cells in worksheet.iter_rows():
            number += 1
            for cell in cells:
                actual = getattr(cell, "row", None)
                if actual:
                    number = int(actual)
                    break
            yield [cell.value for cell in cells], number

    def _is_header_row(self, values: list[Any], texts: list[str]) -> bool:
        """Czy wiersz wyglada na naglowek kolumn: wiekszosc komorek to tekst."""
        filled = [index for index, text in enumerate(texts) if text]
        if not filled:
            return False
        textual = sum(1 for index in filled if isinstance(values[index], str))
        return textual * 2 > len(filled)

    def _row_text(self, texts: list[str], header: list[str]) -> str:
        """Sklada tekst wiersza, opisujac wartosci nazwami kolumn gdy sa znane."""
        parts: list[str] = []
        for index, text in enumerate(texts):
            if not text:
                continue
            label = header[index] if index < len(header) else ""
            parts.append(f"{label}: {text}" if label else text)
        return _CELL_SEPARATOR.join(parts)

    def _read_sheet(
        self,
        worksheet: Worksheet,
        context: ExtractionContext,
        result: ExtractionResult,
        order: int,
    ) -> _SheetOutcome:
        """Dodaje sekcje jednego arkusza do wyniku."""
        title = str(worksheet.title or "Arkusz")
        sheet_text = clean_text(f"Arkusz: {title}")
        sheet_index = len(result.sections)
        if sheet_text:
            result.sections.append(
                ExtractedSection(
                    text=sheet_text,
                    kind="sheet",
                    order=order,
                    sheet=title,
                    extra=self._sheet_extra(worksheet),
                )
            )
            order += 1

        limit = max(1, int(context.sheet_max_rows))
        header: list[str] = []
        heading: str | None = None
        content_rows = 0
        used_rows = 0
        scanned = 0
        first_filled_row = True
        failed = False

        try:
            for values, row_number in self._iter_rows(worksheet):
                scanned += 1
                if scanned % _CHECKPOINT_EVERY == 0:
                    context.checkpoint()
                texts = [format_cell(value) for value in values]
                if not any(texts):
                    continue
                if used_rows >= limit:
                    result.warnings.append(
                        f"Arkusz '{title}': przekroczono limit {limit} wierszy, "
                        "dalsza część arkusza została pominięta."
                    )
                    break
                used_rows += 1

                if first_filled_row:
                    first_filled_row = False
                    if self._is_header_row(values, texts):
                        header = texts
                        heading = clean_text(_CELL_SEPARATOR.join(text for text in texts if text))
                        if heading:
                            result.sections.append(
                                ExtractedSection(
                                    text=heading,
                                    kind="table_header",
                                    order=order,
                                    sheet=title,
                                    row=row_number,
                                    heading=heading,
                                )
                            )
                            order += 1
                            content_rows += 1
                        continue

                text = clean_text(self._row_text(texts, header))
                if not text:
                    continue
                result.sections.append(
                    ExtractedSection(
                        text=text,
                        kind="table_row",
                        order=order,
                        sheet=title,
                        row=row_number,
                        heading=heading,
                    )
                )
                order += 1
                content_rows += 1
        except FindDocsError:
            raise
        except Exception as exc:
            failed = True
            result.warnings.append(
                f"Arkusz '{title}': odczyt przerwano po błędzie ({type(exc).__name__})."
            )

        if content_rows == 0 and sheet_text:
            del result.sections[sheet_index]
            order -= 1
        return _SheetOutcome(next_order=order, rows=content_rows, failed=failed)

    def _sheet_extra(self, worksheet: Worksheet) -> dict[str, Any]:
        """Dodatkowe informacje o arkuszu, np. czy jest ukryty."""
        state = str(getattr(worksheet, "sheet_state", "visible") or "visible")
        extra: dict[str, Any] = {}
        if state != "visible":
            extra["ukryty"] = True
            extra["stan_arkusza"] = state
        return extra


__all__ = ["XlsxExtractor", "format_cell"]
